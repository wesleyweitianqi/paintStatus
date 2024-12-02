from openpyxl import load_workbook 
import json
import sys

def add_data_to_workbook(data):

  file_path = '133form.xlsx'

  try: 
    wb = load_workbook(file_path)
    ws = wb.sheetnames
    print(ws[1])
    sheet = wb[ws[1]]
  except Exception:
    print("Error loading the Excel file")


  for row in data:
    sheet.append(row);

  wb.save('output.xlsx')

  print('files saved successfulli')

if __name__ == '__main__':
  try:
    input_data = sys.stdin.read()
    data = json.loads(input_data)
    add_data_to_workbook(data)
  except Exception as e:
    print(f"Error adding data to the Excel file: {e}")


